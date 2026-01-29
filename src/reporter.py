from __future__ import annotations

"""
reporter.py
-----------
TASK 2: Excel report generator

Input:
- moved_files list (list of dicts) from organizer run log JSON

Output:
- reports/report_YYYY-MM-DD.xlsx

Excel content:
1) "Moved Files" sheet: all file move details
2) "Summary" sheet: files per category + total MB
3) "Top 10 Largest" sheet: biggest files
Plus:
- bar chart (files moved by category) on Summary sheet
"""

from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference


def generate_excel_report(moved_files: List[Dict[str, Any]], report_path: Path) -> Path:
    """
    Create Excel report from moved_files list.
    """
    report_path.parent.mkdir(parents=True, exist_ok=True)

    # Convert moved_files (list of dicts) into a table (DataFrame)
    df = pd.DataFrame(moved_files)

    # If no files moved, keep report valid
    if df.empty:
        df = pd.DataFrame([{
            "src": "",
            "dst": "",
            "category": "No files moved",
            "size_bytes": 0,
            "moved_at": "",
            "dry_run": True
        }])

    # Summary: file count + total bytes per category
    summary = (
        df.groupby("category")
          .agg(files=("category", "count"), total_bytes=("size_bytes", "sum"))
          .reset_index()
          .sort_values("files", ascending=False)
    )
    summary["total_mb"] = (summary["total_bytes"] / (1024 * 1024)).round(2)

    # Top 10 largest files
    top10 = df.sort_values("size_bytes", ascending=False).head(10).copy()
    top10["size_mb"] = (top10["size_bytes"] / (1024 * 1024)).round(2)

    # Write sheets using pandas
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Moved Files")
        summary.to_excel(writer, index=False, sheet_name="Summary")
        top10.to_excel(writer, index=False, sheet_name="Top 10 Largest")

    # Add chart using openpyxl 
    wb = load_workbook(report_path)
    ws = wb["Summary"]

    # Create a bar chart: Category vs Files
    chart = BarChart()
    chart.title = "Files moved by category"
    chart.y_axis.title = "Files"
    chart.x_axis.title = "Category"

    # Data: "files" column is 2nd column in Summary sheet
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Place chart at cell E2
    ws.add_chart(chart, "E2")

    wb.save(report_path)
    return report_path
